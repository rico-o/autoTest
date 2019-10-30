#coding=utf-8
from openpyxl import load_workbook
import os
class ExcelUtil(object):
 def __init__(self,excel_path=None,sheet_name=None):
      """获取 excel 工作表"""
      if excel_path == None:
         # self.excel_path='C:\\Users\jsb05\Desktop\\bug\\casedata.xlsx'
         current_path = os.path.abspath(os.path.dirname(__file__))
         self.excel_path = current_path + '/../data/casedata.xlsx'

      else:
          self.excel_path = excel_path
      if sheet_name == None:
          self.sheet_name = "Sheet"
      else:
          self.sheet_name = sheet_name
      # 打开工作表
          self.data = load_workbook(self.excel_path)
          self.sheet = self.data[self.sheet_name]
 def get_data(self):
  """
获取文件数据
每一行数据一个list，所有的数据一个大list
"""
  rows = self.sheet.rows
  row_num = self.sheet.max_row    #行
  col_num = self.sheet.max_column #列
  if row_num <= 1:
   print("总行数小于1,没有数据")
  else:
    case_all = []
    for row in rows:
     case = []
     for i in range(col_num):
      case.append(row[i].value)
     case_all.append(case)
    return case_all

class Worker_ExcelUtil(object):
 def __init__(self,excel_path=None,sheet_name=None):
      """获取 excel 工作表"""
      if excel_path == None:
         current_path = os.path.abspath(os.path.dirname(__file__))
         self.excel_path = current_path + '/../data/worker_casedata.xlsx'

      else:
          self.excel_path = excel_path
      if sheet_name == None:
          self.sheet_name = "Sheet"
      else:
          self.sheet_name = sheet_name
      # 打开工作表
          self.data = load_workbook(self.excel_path)
          self.sheet = self.data[self.sheet_name]
 def get_data(self):
  """
获取文件数据
每一行数据一个list，所有的数据一个大list
"""
  rows = self.sheet.rows
  row_num = self.sheet.max_row    #行
  col_num = self.sheet.max_column #列
  if row_num <= 1:
   print("总行数小于1,没有数据")
  else:
    case_all = []
    for row in rows:
     case = []
     for i in range(col_num):
      case.append(row[i].value)
     case_all.append(case)
    return case_all

class Serach_ExcelUtil(object):
 def __init__(self,excel_path=None,sheet_name=None):
      """获取 excel 工作表"""
      if excel_path == None:
         current_path = os.path.abspath(os.path.dirname(__file__))
         self.excel_path = current_path + '/../data/search_casedata.xlsx'

      else:
          self.excel_path = excel_path
      if sheet_name == None:
          self.sheet_name = "Sheet"
      else:
          self.sheet_name = sheet_name
      # 打开工作表
          self.data = load_workbook(self.excel_path)
          self.sheet = self.data[self.sheet_name]
 def get_data(self):
  """
获取文件数据
每一行数据一个list，所有的数据一个大list
"""
  rows = self.sheet.rows
  row_num = self.sheet.max_row    #行
  col_num = self.sheet.max_column #列
  if row_num <= 1:
   print("总行数小于1,没有数据")
  else:
    case_all = []
    for row in rows:
     case = []
     for i in range(col_num):
      case.append(row[i].value)
     case_all.append(case)
    return case_all
# if __name__ == '__main__':
#     sheet = 'player'
#     file = ExcelUtil(sheet_name=sheet)
#     print(file.get_data())
#     sheet = 'worker'
#     file =  Worker_ExcelUtil(sheet_name=sheet)
#     print(file.get_data())
#     sheet = 'player'
#     file = Serach_ExcelUtil(sheet_name=sheet)
#     print(file.get_data())

